"""Document views — list, detail, PDF generation via WeasyPrint."""

from datetime import date, timedelta
from decimal import ROUND_HALF_UP, Decimal

from django.contrib.auth.decorators import login_required
from django.http import HttpResponse
from django.shortcuts import get_object_or_404, redirect, render
from django.template.loader import render_to_string

from accounts.mixins import role_required
from accounts.models import Role
from jobs.models import Job, JobStatus, PaymentStatus

from .models import Document, DocumentItem, DocumentType, Setting


@login_required
def document_list(request):
    doc_type = request.GET.get("type", "")
    docs = Document.objects.filter(is_void=False).select_related("job__customer").order_by("-issued_at")
    if doc_type:
        docs = docs.filter(document_type=doc_type)
    return render(
        request,
        "documents/list.html",
        {"documents": docs, "doc_type": doc_type, "document_types": DocumentType.choices},
    )


@login_required
def document_detail(request, pk):
    doc = get_object_or_404(Document.objects.select_related("job__customer", "issued_by"), pk=pk)
    return render(request, "documents/detail.html", {"document": doc})


@login_required
def document_pdf(request, pk):
    """Generate PDF via WeasyPrint and stream to browser."""
    from weasyprint import HTML

    doc = get_object_or_404(Document.objects.prefetch_related("items").select_related("issued_by"), pk=pk)
    shop_info = {
        "name": Setting.get("shop_name", "ร้านพิมพ์"),
        "address": Setting.get("shop_address", ""),
        "tax_id": Setting.get("shop_tax_id", ""),
        "phone": Setting.get("shop_phone", ""),
    }
    html_string = render_to_string(
        "documents/pdf/document.html",
        {"document": doc, "shop": shop_info},
        request=request,
    )
    pdf_bytes = HTML(string=html_string, base_url=request.build_absolute_uri("/")).write_pdf()
    response = HttpResponse(pdf_bytes, content_type="application/pdf")
    response["Content-Disposition"] = f'inline; filename="{doc.document_number}.pdf"'
    return response


@role_required(Role.COUNTER, Role.OWNER)
def create_quotation(request, job_id):
    """Create a quotation document for a job."""
    job = get_object_or_404(Job.objects.select_related("customer"), pk=job_id)

    doc = Document.objects.create(
        job=job,
        document_type=DocumentType.QUOTATION,
        customer_name=job.customer.name,
        customer_address=job.customer.billing_address,
        customer_tax_id=job.customer.tax_id,
        subtotal=job.quoted_price - job.discount_amount,
        vat_rate=0,
        vat_amount=0,
        total_amount=job.quoted_price - job.discount_amount,
        issued_by=request.user,
    )
    DocumentItem.objects.create(
        document=doc,
        description=f"{job.product_type.name} — {job.title}",
        quantity=job.quantity,
        unit=job.product_type.unit,
        unit_price=(job.quoted_price - job.discount_amount) / max(job.quantity, 1),
    )
    return redirect("documents:pdf", pk=doc.pk)


def get_or_create_receipt(job, created_by):
    """
    Get existing receipt for job or create one.
    Called from payments.views after payment recorded.
    """
    existing = Document.objects.filter(
        job=job,
        document_type=DocumentType.RECEIPT,
        is_void=False,
    ).first()
    if existing:
        return existing

    total = job.quoted_price - job.discount_amount
    doc = Document.objects.create(
        job=job,
        document_type=DocumentType.RECEIPT,
        customer_name=job.customer.name,
        customer_address=job.customer.billing_address,
        customer_tax_id=job.customer.tax_id,
        subtotal=total,
        vat_rate=0,
        vat_amount=0,
        total_amount=total,
        issued_by=created_by,
    )
    DocumentItem.objects.create(
        document=doc,
        description=f"{job.product_type.name} — {job.title}",
        quantity=job.quantity,
        unit=job.product_type.unit,
        unit_price=total / max(job.quantity, 1),
    )
    return doc


@role_required(Role.COUNTER, Role.OWNER)
def create_tax_invoice(request, job_id):
    """Create a VAT-inclusive tax invoice for a job. POST only."""
    job = get_object_or_404(Job.objects.select_related("customer"), pk=job_id)

    # Guard: return existing non-void tax invoice if one already exists
    existing = Document.objects.filter(
        job=job,
        document_type=DocumentType.TAX_INVOICE,
        is_void=False,
    ).first()
    if existing:
        return redirect("documents:pdf", pk=existing.pk)

    vat_rate = Decimal(Setting.get("vat_rate", "7"))
    total = job.quoted_price - job.discount_amount
    divisor = 1 + vat_rate / Decimal("100")
    subtotal = (total / divisor).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)
    vat_amount = total - subtotal

    doc = Document.objects.create(
        job=job,
        document_type=DocumentType.TAX_INVOICE,
        customer_name=job.customer.name,
        customer_address=job.customer.billing_address,
        customer_tax_id=job.customer.tax_id,
        subtotal=subtotal,
        vat_rate=vat_rate,
        vat_amount=vat_amount,
        total_amount=total,
        issued_by=request.user,
    )
    DocumentItem.objects.create(
        document=doc,
        description=f"{job.product_type.name} — {job.title}",
        quantity=job.quantity,
        unit=job.product_type.unit,
        unit_price=(subtotal / max(job.quantity, 1)).quantize(
            Decimal("0.01"), rounding=ROUND_HALF_UP
        ),
    )
    return redirect("documents:pdf", pk=doc.pk)


@role_required(Role.OWNER)
def void_document(request, pk):
    """Void a document. POST only. Owner only."""
    if request.method != "POST":
        return redirect("documents:detail", pk=pk)
    doc = get_object_or_404(Document, pk=pk)
    doc.is_void = True
    doc.save(update_fields=["is_void"])
    return redirect("documents:detail", pk=pk)


@login_required
def aging_report(request):
    """Accounts-receivable aging report — unpaid/partial jobs by overdue bucket."""
    billable_statuses = [
        JobStatus.APPROVED,
        JobStatus.PRINTING,
        JobStatus.CUTTING,
        JobStatus.LAMINATING,
        JobStatus.READY,
        JobStatus.COMPLETED,
    ]
    jobs = (
        Job.objects.filter(
            payment_status__in=[PaymentStatus.UNPAID, PaymentStatus.PARTIAL],
            status__in=billable_statuses,
        )
        .select_related("customer__customer_type")
        .prefetch_related("documents")
        .order_by("created_at")
    )

    today = date.today()
    rows = []
    bucket_totals = {"current": Decimal("0"), "1_30": Decimal("0"), "31_60": Decimal("0"), "61_90": Decimal("0"), "91_plus": Decimal("0")}

    for job in jobs:
        invoice_doc = next(
            (
                d
                for d in job.documents.all()
                if d.document_type in (DocumentType.TAX_INVOICE, DocumentType.RECEIPT)
                and not d.is_void
            ),
            None,
        )
        invoice_date = invoice_doc.issued_at.date() if invoice_doc else job.created_at.date()
        credit_days = getattr(getattr(job.customer, "customer_type", None), "credit_days", 0) or 0
        due_date = invoice_date + timedelta(days=credit_days)
        days_overdue = max(0, (today - due_date).days)
        balance = job.quoted_price - job.discount_amount - job.total_paid

        if days_overdue == 0:
            bucket = "current"
        elif days_overdue <= 30:
            bucket = "1_30"
        elif days_overdue <= 60:
            bucket = "31_60"
        elif days_overdue <= 90:
            bucket = "61_90"
        else:
            bucket = "91_plus"

        bucket_totals[bucket] += balance
        rows.append(
            {
                "job": job,
                "invoice_doc": invoice_doc,
                "invoice_date": invoice_date,
                "due_date": due_date,
                "days_overdue": days_overdue,
                "balance": balance,
                "bucket": bucket,
            }
        )

    grand_total = sum(bucket_totals.values())
    return render(
        request,
        "documents/aging.html",
        {
            "rows": rows,
            "bucket_totals": bucket_totals,
            "grand_total": grand_total,
            "today": today,
        },
    )


@login_required
def aging_export_excel(request):
    """Export aging report as Excel (.xlsx)."""
    import io
    from datetime import date as date_cls

    import openpyxl
    from openpyxl.styles import Alignment, Font, PatternFill

    # --- reuse the aging queryset logic ---
    billable_statuses = [
        JobStatus.APPROVED,
        JobStatus.PRINTING,
        JobStatus.CUTTING,
        JobStatus.LAMINATING,
        JobStatus.READY,
        JobStatus.COMPLETED,
    ]
    jobs = (
        Job.objects.filter(
            payment_status__in=[PaymentStatus.UNPAID, PaymentStatus.PARTIAL],
            status__in=billable_statuses,
        )
        .select_related("customer__customer_type")
        .prefetch_related("documents")
        .order_by("created_at")
    )

    today = date.today()
    rows = []
    bucket_totals = {
        "current": Decimal("0"), "1_30": Decimal("0"),
        "31_60": Decimal("0"), "61_90": Decimal("0"), "91_plus": Decimal("0"),
    }
    for job in jobs:
        invoice_doc = next(
            (d for d in job.documents.all()
             if d.document_type in (DocumentType.TAX_INVOICE, DocumentType.RECEIPT) and not d.is_void),
            None,
        )
        invoice_date = invoice_doc.issued_at.date() if invoice_doc else job.created_at.date()
        credit_days = getattr(getattr(job.customer, "customer_type", None), "credit_days", 0) or 0
        due_date = invoice_date + timedelta(days=credit_days)
        days_overdue = max(0, (today - due_date).days)
        balance = job.quoted_price - job.discount_amount - job.total_paid
        if days_overdue == 0:
            bucket = "current"
        elif days_overdue <= 30:
            bucket = "1_30"
        elif days_overdue <= 60:
            bucket = "31_60"
        elif days_overdue <= 90:
            bucket = "61_90"
        else:
            bucket = "91_plus"
        bucket_totals[bucket] += balance
        rows.append({
            "job": job, "invoice_doc": invoice_doc, "invoice_date": invoice_date,
            "due_date": due_date, "days_overdue": days_overdue, "balance": balance, "bucket": bucket,
        })

    bucket_labels = {
        "current": "ยังไม่เกินกำหนด", "1_30": "1–30 วัน",
        "31_60": "31–60 วัน", "61_90": "61–90 วัน", "91_plus": "91+ วัน",
    }

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Aging Report"

    header_fill = PatternFill("solid", fgColor="4F46E5")
    header_font = Font(bold=True, color="FFFFFF", size=10)
    headers = ["#", "ลูกค้า", "งาน", "วันที่เอกสาร", "วันครบกำหนด", "เกินกำหนด (วัน)", "ยอดค้าง", "กลุ่ม"]
    ws.append(headers)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    for i, row in enumerate(rows, 1):
        ws.append([
            i,
            row["job"].customer.name,
            row["job"].title,
            row["invoice_date"],
            row["due_date"],
            row["days_overdue"],
            float(row["balance"]),
            bucket_labels.get(row["bucket"], row["bucket"]),
        ])

    ws.append([])
    ws.append(["", "", "", "", "", "รวม", float(sum(bucket_totals.values())), ""])
    for key, label in bucket_labels.items():
        ws.append(["", "", "", "", "", label, float(bucket_totals[key]), ""])

    for col in ws.columns:
        max_len = max((len(str(cell.value)) if cell.value else 0) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 40)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    filename = f"aging_report_{today.strftime('%Y%m%d')}.xlsx"
    response = HttpResponse(
        buf.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    return response


@login_required
def vat_report_export_excel(request):
    """Export VAT report for a given month/year as Excel (.xlsx)."""
    import io

    import openpyxl
    from openpyxl.styles import Alignment, Font, PatternFill

    today = date.today()
    try:
        month = int(request.GET.get("month", today.month))
        year = int(request.GET.get("year", today.year))
    except (ValueError, TypeError):
        month, year = today.month, today.year
    month = max(1, min(12, month))

    docs = (
        Document.objects.filter(
            document_type=DocumentType.TAX_INVOICE,
            vat_rate__gt=0,
            is_void=False,
            issued_at__year=year,
            issued_at__month=month,
        )
        .select_related("job__customer")
        .order_by("issued_at")
    )

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"VAT {month:02d}-{year}"

    header_fill = PatternFill("solid", fgColor="4F46E5")
    header_font = Font(bold=True, color="FFFFFF", size=10)
    headers = ["ลำดับ", "เลขที่เอกสาร", "วันที่", "ชื่อลูกค้า", "เลขผู้เสียภาษี", "ยอดก่อน VAT", "VAT", "ยอดรวม"]
    ws.append(headers)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    total_subtotal = Decimal("0")
    total_vat = Decimal("0")
    total_amount = Decimal("0")

    for i, doc in enumerate(docs, 1):
        ws.append([
            i,
            doc.document_number,
            doc.issued_at.date(),
            doc.customer_name,
            doc.customer_tax_id or "",
            float(doc.subtotal),
            float(doc.vat_amount),
            float(doc.total_amount),
        ])
        total_subtotal += doc.subtotal
        total_vat += doc.vat_amount
        total_amount += doc.total_amount

    ws.append([])
    ws.append(["", "", "", "รวม", "", float(total_subtotal), float(total_vat), float(total_amount)])
    last_row = ws.max_row
    summary_font = Font(bold=True)
    for cell in ws[last_row]:
        cell.font = summary_font

    for col in ws.columns:
        max_len = max((len(str(cell.value)) if cell.value else 0) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 40)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    filename = f"vat_report_{month:02d}_{year}.xlsx"
    response = HttpResponse(
        buf.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    return response


@role_required(Role.OWNER, Role.ACCOUNTANT)
def revenue_report(request):
    """Monthly revenue summary — payments received, by product type and method."""
    today = date.today()
    try:
        month = int(request.GET.get("month", today.month))
        year = int(request.GET.get("year", today.year))
    except (ValueError, TypeError):
        month, year = today.month, today.year
    month = max(1, min(12, month))

    from django.db.models import Count, F, Sum

    from payments.models import Payment

    payments = Payment.objects.filter(
        received_at__year=year,
        received_at__month=month,
    ).select_related("job__product_type")

    total = payments.aggregate(total=Sum(F("amount") + F("wht_amount")))["total"] or 0

    by_product = (
        payments.values("job__product_type__name")
        .annotate(total=Sum(F("amount") + F("wht_amount")), count=Count("id"))
        .order_by("-total")
    )

    by_method = (
        payments.values("method")
        .annotate(total=Sum(F("amount") + F("wht_amount")))
        .order_by("-total")
    )

    completed_count = Job.objects.filter(
        status=JobStatus.COMPLETED,
        updated_at__year=year,
        updated_at__month=month,
    ).count()

    if month == 1:
        prev_month, prev_year = 12, year - 1
    else:
        prev_month, prev_year = month - 1, year

    if month == 12:
        next_month, next_year = 1, year + 1
    else:
        next_month, next_year = month + 1, year

    thai_months = [
        "", "มกราคม", "กุมภาพันธ์", "มีนาคม", "เมษายน",
        "พฤษภาคม", "มิถุนายน", "กรกฎาคม", "สิงหาคม",
        "กันยายน", "ตุลาคม", "พฤศจิกายน", "ธันวาคม",
    ]

    return render(request, "documents/revenue.html", {
        "total": total,
        "by_product": by_product,
        "by_method": by_method,
        "completed_count": completed_count,
        "month": month,
        "month_name": thai_months[month],
        "year": year,
        "prev_month": prev_month,
        "prev_year": prev_year,
        "next_month": next_month,
        "next_year": next_year,
    })


@login_required
def vat_report(request):
    """Monthly VAT output tax report."""
    today = date.today()
    try:
        month = int(request.GET.get("month", today.month))
        year = int(request.GET.get("year", today.year))
    except (ValueError, TypeError):
        month, year = today.month, today.year

    # Clamp month to 1-12
    month = max(1, min(12, month))

    docs = (
        Document.objects.filter(
            document_type=DocumentType.TAX_INVOICE,
            vat_rate__gt=0,
            is_void=False,
            issued_at__year=year,
            issued_at__month=month,
        )
        .select_related("job__customer")
        .order_by("issued_at")
    )

    from django.db.models import Sum

    totals = docs.aggregate(
        total_subtotal=Sum("subtotal"),
        total_vat=Sum("vat_amount"),
        total_amount=Sum("total_amount"),
    )

    # Prev / next month navigation
    if month == 1:
        prev_month, prev_year = 12, year - 1
    else:
        prev_month, prev_year = month - 1, year

    if month == 12:
        next_month, next_year = 1, year + 1
    else:
        next_month, next_year = month + 1, year

    thai_months = [
        "", "มกราคม", "กุมภาพันธ์", "มีนาคม", "เมษายน",
        "พฤษภาคม", "มิถุนายน", "กรกฎาคม", "สิงหาคม",
        "กันยายน", "ตุลาคม", "พฤศจิกายน", "ธันวาคม",
    ]
    return render(
        request,
        "documents/vat_report.html",
        {
            "docs": docs,
            "totals": totals,
            "month": month,
            "month_name": thai_months[month],
            "year": year,
            "prev_month": prev_month,
            "prev_year": prev_year,
            "next_month": next_month,
            "next_year": next_year,
        },
    )


@role_required(Role.COUNTER, Role.OWNER, Role.ACCOUNTANT)
def monthly_statement(request, customer_id):
    """Generate a monthly statement PDF for a customer."""
    from weasyprint import HTML

    from customers.models import Customer

    from django.db.models import Sum

    customer = get_object_or_404(Customer, pk=customer_id)
    today = date.today()
    try:
        month = int(request.GET.get("month", today.month))
        year = int(request.GET.get("year", today.year))
    except (ValueError, TypeError):
        month, year = today.month, today.year
    month = max(1, min(12, month))

    docs = Document.objects.filter(
        job__customer=customer,
        is_void=False,
        issued_at__year=year,
        issued_at__month=month,
        document_type__in=[DocumentType.TAX_INVOICE, DocumentType.RECEIPT],
    ).select_related("job").order_by("issued_at")

    totals = docs.aggregate(total=Sum("total_amount"), vat=Sum("vat_amount"))
    shop_info = {k: Setting.get(k, "") for k in ["shop_name", "shop_address", "shop_tax_id", "shop_phone"]}

    thai_months = [
        "", "มกราคม", "กุมภาพันธ์", "มีนาคม", "เมษายน",
        "พฤษภาคม", "มิถุนายน", "กรกฎาคม", "สิงหาคม",
        "กันยายน", "ตุลาคม", "พฤศจิกายน", "ธันวาคม",
    ]

    html_string = render_to_string(
        "documents/pdf/statement.html",
        {
            "customer": customer,
            "docs": docs,
            "totals": totals,
            "month": month,
            "month_name": thai_months[month],
            "year": year,
            "shop": shop_info,
        },
        request=request,
    )
    pdf = HTML(string=html_string, base_url=request.build_absolute_uri("/")).write_pdf()
    response = HttpResponse(pdf, content_type="application/pdf")
    response["Content-Disposition"] = f'inline; filename="statement_{customer_id}_{year}_{month:02d}.pdf"'
    return response
